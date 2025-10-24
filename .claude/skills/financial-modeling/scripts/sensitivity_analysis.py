#!/usr/bin/env python3
"""
ABOUTME: Sensitivity and scenario analysis generation
ABOUTME: Creates sensitivity tables, tornado charts, and scenario models

Generates sensitivity tables and scenario analysis for financial models
to analyze output sensitivity to input assumption changes.
"""

from typing import Dict, List, Tuple, Callable, Optional
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class SensitivityAnalyzer:
    """Performs sensitivity analysis on financial models."""
    
    def __init__(self):
        self.wb = Workbook()
        self.wb.remove(self.wb.active)
    
    def create_one_way_sensitivity(
        self,
        base_case_value: float,
        variable_name: str,
        variable_ranges: List[float],
        calculation_function: Callable[[float], float],
        output_name: str = "Output"
    ) -> Dict[float, float]:
        """
        Create a one-way sensitivity table.
        
        Args:
            base_case_value: Base case value for the variable
            variable_name: Name of the variable being sensitized
            variable_ranges: List of values to test (typically -30%, -20%, -10%, 0%, +10%, +20%, +30%)
            calculation_function: Function that takes variable value and returns output
            output_name: Name of the output metric
        
        Returns:
            Dictionary mapping variable values to output values
        """
        results = {}
        for var_value in variable_ranges:
            output = calculation_function(var_value)
            results[var_value] = output
        
        return results
    
    def create_two_way_sensitivity(
        self,
        base_case_values: Tuple[float, float],
        variable_names: Tuple[str, str],
        variable_ranges: Tuple[List[float], List[float]],
        calculation_function: Callable[[float, float], float],
        output_name: str = "Output"
    ) -> Dict[Tuple[float, float], float]:
        """
        Create a two-way sensitivity table.
        
        Args:
            base_case_values: Tuple of base case values for both variables
            variable_names: Tuple of variable names
            variable_ranges: Tuple of lists containing ranges for each variable
            calculation_function: Function that takes two variable values and returns output
            output_name: Name of the output metric
        
        Returns:
            Dictionary mapping (var1, var2) tuples to output values
        """
        results = {}
        for var1_value in variable_ranges[0]:
            for var2_value in variable_ranges[1]:
                output = calculation_function(var1_value, var2_value)
                results[(var1_value, var2_value)] = output
        
        return results
    
    def calculate_sensitivity_range(
        self,
        base_case: float,
        percentages: List[float] = None
    ) -> List[float]:
        """
        Calculate sensitivity range values based on base case and percentages.
        
        Args:
            base_case: Base case value
            percentages: List of percentage changes (default: -30%, -20%, -10%, 0%, 10%, 20%, 30%)
        
        Returns:
            List of calculated values
        """
        if percentages is None:
            percentages = [-0.30, -0.20, -0.10, 0, 0.10, 0.20, 0.30]
        
        return [base_case * (1 + pct) for pct in percentages]
    
    def create_scenario_analysis(
        self,
        scenarios: Dict[str, Dict[str, float]],
        calculation_function: Callable[[Dict[str, float]], float],
        output_name: str = "Output"
    ) -> Dict[str, float]:
        """
        Create scenario analysis with multiple assumptions.
        
        Args:
            scenarios: Dictionary of scenario names to assumption dictionaries
            calculation_function: Function that takes assumptions dict and returns output
            output_name: Name of the output metric
        
        Returns:
            Dictionary mapping scenario names to output values
        """
        results = {}
        for scenario_name, assumptions in scenarios.items():
            output = calculation_function(assumptions)
            results[scenario_name] = output
        
        return results
    
    def add_sensitivity_sheet(
        self,
        sheet_name: str,
        base_value: float,
        variable_name: str,
        results: Dict[float, float],
        output_name: str = "Valuation"
    ) -> None:
        """Add a one-way sensitivity table to a worksheet."""
        ws = self.wb.create_sheet(sheet_name)
        
        # Header
        ws['A1'] = f"One-Way Sensitivity Analysis: {variable_name}"
        ws['A1'].font = Font(bold=True, size=12, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        ws.merge_cells('A1:B1')
        
        # Column headers
        ws['A3'] = f"{variable_name} Value"
        ws['B3'] = output_name
        for cell in ['A3', 'B3']:
            ws[cell].font = Font(bold=True, color="FFFFFF")
            ws[cell].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        
        # Data rows
        row = 4
        for var_value in sorted(results.keys()):
            output_value = results[var_value]
            
            ws.cell(row=row, column=1).value = var_value
            ws.cell(row=row, column=2).value = output_value
            
            # Highlight base case
            if abs(var_value - base_value) < 0.01:
                ws.cell(row=row, column=1).fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                ws.cell(row=row, column=2).fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            
            row += 1
        
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 20
    
    def add_two_way_sensitivity_sheet(
        self,
        sheet_name: str,
        var1_name: str,
        var2_name: str,
        var1_values: List[float],
        var2_values: List[float],
        results: Dict[Tuple[float, float], float],
        output_name: str = "Valuation"
    ) -> None:
        """Add a two-way sensitivity table to a worksheet."""
        ws = self.wb.create_sheet(sheet_name)
        
        # Header
        ws['A1'] = f"Two-Way Sensitivity Analysis: {var1_name} vs {var2_name}"
        ws['A1'].font = Font(bold=True, size=12, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        ws.merge_cells(f'A1:{get_column_letter(len(var2_values) + 1)}1')
        
        # Headers
        ws['A3'] = var1_name
        ws['A3'].font = Font(bold=True, color="FFFFFF")
        ws['A3'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        
        for col_idx, var2_value in enumerate(var2_values, start=2):
            cell = ws.cell(row=3, column=col_idx)
            cell.value = var2_value
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        
        # Data
        for row_idx, var1_value in enumerate(var1_values, start=4):
            ws.cell(row=row_idx, column=1).value = var1_value
            ws.cell(row=row_idx, column=1).font = Font(bold=True)
            ws.cell(row=row_idx, column=1).fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            
            for col_idx, var2_value in enumerate(var2_values, start=2):
                key = (var1_value, var2_value)
                if key in results:
                    ws.cell(row=row_idx, column=col_idx).value = results[key]
        
        # Set column widths
        ws.column_dimensions['A'].width = 15
        for col in range(2, len(var2_values) + 2):
            ws.column_dimensions[get_column_letter(col)].width = 15
    
    def add_scenario_sheet(
        self,
        sheet_name: str,
        scenarios: Dict[str, float],
        assumptions: Dict[str, Dict[str, float]],
        output_name: str = "Valuation"
    ) -> None:
        """Add scenario analysis sheet to workbook."""
        ws = self.wb.create_sheet(sheet_name)
        
        # Header
        ws['A1'] = "Scenario Analysis"
        ws['A1'].font = Font(bold=True, size=12, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        ws.merge_cells('A1:D1')
        
        # Scenario summary
        ws['A3'] = 'Scenario'
        ws['B3'] = output_name
        ws['C3'] = 'vs Base Case'
        ws['D3'] = '% Change'
        
        for cell in ['A3', 'B3', 'C3', 'D3']:
            ws[cell].font = Font(bold=True, color="FFFFFF")
            ws[cell].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        
        base_value = scenarios.get('Base Case', 0)
        row = 4
        
        for scenario_name, output_value in scenarios.items():
            ws.cell(row=row, column=1).value = scenario_name
            ws.cell(row=row, column=2).value = output_value
            ws.cell(row=row, column=3).value = output_value - base_value
            
            if base_value != 0:
                ws.cell(row=row, column=4).value = (output_value - base_value) / base_value
                ws.cell(row=row, column=4).number_format = '0.0%'
            
            row += 1
        
        # Assumptions detail
        if assumptions:
            row += 2
            ws.cell(row=row, column=1).value = "Scenario Assumptions"
            ws.cell(row=row, column=1).font = Font(bold=True, size=11)
            row += 1
            
            for scenario_name, scenario_assumptions in assumptions.items():
                ws.cell(row=row, column=1).value = scenario_name
                ws.cell(row=row, column=1).font = Font(bold=True)
                ws.cell(row=row, column=1).fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
                row += 1
                
                for assumption_name, value in scenario_assumptions.items():
                    ws.cell(row=row, column=2).value = assumption_name
                    ws.cell(row=row, column=3).value = value
                    row += 1
                
                row += 1
        
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
    
    def save(self, filename: str) -> None:
        """Save the workbook."""
        self.wb.save(filename)
        print(f"[OK] Sensitivity analysis saved to {filename}")


def create_standard_sensitivity_ranges(
    base_value: float,
    percentages: List[float] = None
) -> Tuple[List[float], List[str]]:
    """
    Create standard sensitivity ranges and labels.
    
    Args:
        base_value: Base case value
        percentages: Percentage changes to calculate
    
    Returns:
        Tuple of (values, labels) for sensitivity analysis
    """
    if percentages is None:
        percentages = [-0.30, -0.20, -0.10, 0, 0.10, 0.20, 0.30]
    
    values = [base_value * (1 + pct) for pct in percentages]
    labels = [f"{pct*100:+.0f}%" for pct in percentages]
    
    return values, labels


if __name__ == "__main__":
    print("[INFO] Sensitivity Analysis Module")
    print("[INFO] Use to create sensitivity tables and scenario analysis")
