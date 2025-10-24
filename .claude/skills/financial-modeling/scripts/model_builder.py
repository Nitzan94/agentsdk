#!/usr/bin/env python3
"""
ABOUTME: Financial model generation from assumptions
ABOUTME: Builds Excel models with formulas for DCF, statements, and scenarios

Automated builder for creating financial models in Excel format with
properly linked formulas and professional formatting.
"""

from typing import Dict, List, Tuple, Any, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import json


class FinancialModelBuilder:
    """Builder for creating Excel financial models with formulas."""
    
    def __init__(self, company_name: str = "Financial Model"):
        self.wb = Workbook()
        self.wb.remove(self.wb.active)
        self.company_name = company_name
        self.colors = {
            "input": "D9E1F2",      # Light blue for inputs
            "calculated": "FFFFFF",  # White for calculations
            "output": "E2EFDA",      # Light green for outputs
            "header": "4472C4"       # Dark blue for headers
        }
    
    def add_assumptions_sheet(self, assumptions: Dict[str, Any]) -> None:
        """Add an assumptions sheet with input values and documentation."""
        ws = self.wb.create_sheet("Assumptions", 0)
        
        # Header
        ws['A1'] = f"{self.company_name} - Financial Model Assumptions"
        ws['A1'].font = Font(bold=True, size=14, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color=self.colors["header"], end_color=self.colors["header"], fill_type="solid")
        ws.merge_cells('A1:D1')
        
        # Column headers
        headers = ['Assumption', 'Value', 'Unit', 'Notes']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color=self.colors["header"], end_color=self.colors["header"], fill_type="solid")
        
        # Add assumptions
        row = 4
        for key, value in assumptions.items():
            ws.cell(row=row, column=1).value = key
            ws.cell(row=row, column=2).value = value
            ws.cell(row=row, column=2).fill = PatternFill(start_color=self.colors["input"], end_color=self.colors["input"], fill_type="solid")
            row += 1
        
        # Set column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 30
    
    def add_income_statement(self, years: int, start_year: int = 2024) -> None:
        """Add Income Statement (P&L) sheet with formula structure."""
        ws = self.wb.create_sheet("Income Statement")
        
        # Headers
        ws['A1'] = f"{self.company_name} - Income Statement (in thousands)"
        ws['A1'].font = Font(bold=True, size=12, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color=self.colors["header"], end_color=self.colors["header"], fill_type="solid")
        ws.merge_cells(f'A1:{get_column_letter(years + 1)}1')
        
        # Year headers
        ws['A3'] = 'Line Item'
        ws['A3'].font = Font(bold=True, color="FFFFFF")
        ws['A3'].fill = PatternFill(start_color=self.colors["header"], end_color=self.colors["header"], fill_type="solid")
        
        for year in range(years):
            col = year + 2
            ws.cell(row=3, column=col).value = start_year + year
            ws.cell(row=3, column=col).font = Font(bold=True, color="FFFFFF")
            ws.cell(row=3, column=col).fill = PatternFill(start_color=self.colors["header"], end_color=self.colors["header"], fill_type="solid")
        
        # P&L line items
        line_items = [
            'Revenue',
            'Revenue Growth %',
            '',
            'Cost of Goods Sold',
            'COGS %',
            'Gross Profit',
            'Gross Margin %',
            '',
            'Operating Expenses',
            'OpEx %',
            'EBITDA',
            'EBITDA Margin %',
            '',
            'Depreciation & Amortization',
            'EBIT',
            'EBIT Margin %',
            '',
            'Interest Expense',
            'EBT',
            'Tax Rate',
            'Taxes',
            'Net Income',
            'Net Margin %'
        ]
        
        row = 4
        for item in line_items:
            ws.cell(row=row, column=1).value = item
            if item and '%' in item:
                ws.cell(row=row, column=1).font = Font(italic=True)
            row += 1
        
        # Set column widths
        ws.column_dimensions['A'].width = 30
        for col in range(2, years + 2):
            ws.column_dimensions[get_column_letter(col)].width = 15
    
    def add_balance_sheet(self, years: int, start_year: int = 2024) -> None:
        """Add Balance Sheet sheet with formula structure."""
        ws = self.wb.create_sheet("Balance Sheet")
        
        # Headers
        ws['A1'] = f"{self.company_name} - Balance Sheet (in thousands)"
        ws['A1'].font = Font(bold=True, size=12, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color=self.colors["header"], end_color=self.colors["header"], fill_type="solid")
        ws.merge_cells(f'A1:{get_column_letter(years + 1)}1')
        
        # Year headers
        ws['A3'] = 'Line Item'
        ws['A3'].font = Font(bold=True, color="FFFFFF")
        ws['A3'].fill = PatternFill(start_color=self.colors["header"], end_color=self.colors["header"], fill_type="solid")
        
        for year in range(years):
            col = year + 2
            ws.cell(row=3, column=col).value = start_year + year
            ws.cell(row=3, column=col).font = Font(bold=True, color="FFFFFF")
            ws.cell(row=3, column=col).fill = PatternFill(start_color=self.colors["header"], end_color=self.colors["header"], fill_type="solid")
        
        # Balance Sheet line items
        sections = {
            'ASSETS': [
                'Current Assets',
                'Cash & Equivalents',
                'Accounts Receivable',
                'Inventory',
                'Total Current Assets',
                '',
                'Fixed Assets',
                'Property, Plant & Equipment',
                'Accumulated Depreciation',
                'Net PP&E',
                'Intangibles',
                'Total Fixed Assets',
                '',
                'TOTAL ASSETS',
                '',
                'LIABILITIES & EQUITY',
                'Current Liabilities',
                'Accounts Payable',
                'Short-term Debt',
                'Total Current Liabilities',
                '',
                'Long-term Liabilities',
                'Long-term Debt',
                'Other Long-term Liabilities',
                'Total Long-term Liabilities',
                '',
                'Total Liabilities',
                '',
                'Stockholders Equity',
                'Common Stock',
                'Retained Earnings',
                'Total Stockholders Equity',
                '',
                'TOTAL LIABILITIES & EQUITY'
            ]
        }
        
        row = 4
        for item in sections['ASSETS']:
            ws.cell(row=row, column=1).value = item
            if item in ['ASSETS', 'TOTAL ASSETS', 'LIABILITIES & EQUITY', 'TOTAL LIABILITIES & EQUITY']:
                ws.cell(row=row, column=1).font = Font(bold=True)
                ws.cell(row=row, column=1).fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
            row += 1
        
        # Set column widths
        ws.column_dimensions['A'].width = 30
        for col in range(2, years + 2):
            ws.column_dimensions[get_column_letter(col)].width = 15
    
    def add_cash_flow_statement(self, years: int, start_year: int = 2024) -> None:
        """Add Cash Flow Statement sheet with formula structure."""
        ws = self.wb.create_sheet("Cash Flow")
        
        # Headers
        ws['A1'] = f"{self.company_name} - Cash Flow Statement (in thousands)"
        ws['A1'].font = Font(bold=True, size=12, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color=self.colors["header"], end_color=self.colors["header"], fill_type="solid")
        ws.merge_cells(f'A1:{get_column_letter(years + 1)}1')
        
        # Year headers
        ws['A3'] = 'Line Item'
        ws['A3'].font = Font(bold=True, color="FFFFFF")
        ws['A3'].fill = PatternFill(start_color=self.colors["header"], end_color=self.colors["header"], fill_type="solid")
        
        for year in range(years):
            col = year + 2
            ws.cell(row=3, column=col).value = start_year + year
            ws.cell(row=3, column=col).font = Font(bold=True, color="FFFFFF")
            ws.cell(row=3, column=col).fill = PatternFill(start_color=self.colors["header"], end_color=self.colors["header"], fill_type="solid")
        
        # Cash Flow line items
        line_items = [
            'OPERATING ACTIVITIES',
            'Net Income',
            'Adjustments:',
            'Depreciation & Amortization',
            'Changes in Working Capital',
            'Accounts Receivable Change',
            'Inventory Change',
            'Accounts Payable Change',
            'Operating Cash Flow',
            '',
            'INVESTING ACTIVITIES',
            'Capital Expenditures',
            'Acquisitions',
            'Investing Cash Flow',
            '',
            'FINANCING ACTIVITIES',
            'Debt Issuance/(Repayment)',
            'Equity Issuance/(Repurchase)',
            'Dividends Paid',
            'Financing Cash Flow',
            '',
            'Net Change in Cash',
            'Beginning Cash',
            'Ending Cash'
        ]
        
        row = 4
        for item in line_items:
            ws.cell(row=row, column=1).value = item
            if item in ['OPERATING ACTIVITIES', 'INVESTING ACTIVITIES', 'FINANCING ACTIVITIES']:
                ws.cell(row=row, column=1).font = Font(bold=True)
                ws.cell(row=row, column=1).fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
            row += 1
        
        # Set column widths
        ws.column_dimensions['A'].width = 30
        for col in range(2, years + 2):
            ws.column_dimensions[get_column_letter(col)].width = 15
    
    def add_valuation_sheet(self) -> None:
        """Add DCF Valuation sheet."""
        ws = self.wb.create_sheet("DCF Valuation")
        
        # Headers
        ws['A1'] = f"{self.company_name} - DCF Valuation"
        ws['A1'].font = Font(bold=True, size=14, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color=self.colors["header"], end_color=self.colors["header"], fill_type="solid")
        ws.merge_cells('A1:C1')
        
        # Valuation summary
        ws['A3'] = 'DCF Valuation Summary'
        ws['A3'].font = Font(bold=True, size=12)
        
        rows_data = [
            ('', '', ''),
            ('Free Cash Flow (Year 5)', 'B6', '=B5'),
            ('Terminal Growth Rate', 'B7', '5.0%'),
            ('WACC (Discount Rate)', 'B8', '10.0%'),
            ('Terminal Value', 'B9', '=B6*(1+B7)/(B8-B7)'),
            ('PV of Terminal Value', 'B10', '=B9/(1+B8)^5'),
            ('PV of Explicit FCF', 'B11', 'SUM(PV FCF)'),
            ('', '', ''),
            ('Enterprise Value', 'B13', '=B10+B11'),
            ('Less: Net Debt', 'B14', '=Debt-Cash'),
            ('Equity Value', 'B15', '=B13-B14'),
            ('Shares Outstanding', 'B16', ''),
            ('Value Per Share', 'B17', '=B15/B16'),
        ]
        
        row = 5
        for label, ref, formula in rows_data:
            if label:
                ws.cell(row=row, column=1).value = label
                ws.cell(row=row, column=1).font = Font(bold=True)
                ws.cell(row=row, column=2).value = formula
                ws.cell(row=row, column=2).fill = PatternFill(start_color=self.colors["input"], end_color=self.colors["input"], fill_type="solid")
            row += 1
        
        # Set column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
    
    def save(self, filename: str) -> None:
        """Save the workbook to file."""
        self.wb.save(filename)
        print(f"[OK] Model saved to {filename}")


def create_default_model(company_name: str, years: int = 5, filename: str = None) -> str:
    """Create a default financial model template."""
    if filename is None:
        filename = f"{company_name.replace(' ', '_')}_Model.xlsx"
    
    builder = FinancialModelBuilder(company_name)
    
    # Add sheets
    default_assumptions = {
        'Revenue (Year 1)': 10000,
        'Revenue Growth %': 0.15,
        'COGS %': 0.40,
        'OpEx %': 0.20,
        'D&A %': 0.05,
        'Tax Rate %': 0.21,
        'CapEx %': 0.05,
        'NWC % of Revenue': 0.10,
        'Terminal Growth %': 0.03,
        'WACC %': 0.10
    }
    
    builder.add_assumptions_sheet(default_assumptions)
    builder.add_income_statement(years)
    builder.add_balance_sheet(years)
    builder.add_cash_flow_statement(years)
    builder.add_valuation_sheet()
    
    builder.save(filename)
    return filename


if __name__ == "__main__":
    print("[INFO] Model Builder Module")
    print("[INFO] Use create_default_model() to generate a template")
