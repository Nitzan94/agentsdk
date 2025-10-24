from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
import random

# Create a new workbook
wb = Workbook()

# Remove default sheet
if 'Sheet' in wb.sheetnames:
    wb.remove(wb['Sheet'])

# ===== LEADS SHEET =====
ws_leads = wb.create_sheet("Leads", 0)

# Headers for Leads sheet
leads_headers = [
    "Lead ID", "Company Name", "Contact Name", "Title", "Email", "Phone",
    "Industry", "Country", "Lead Source", "Status", "Priority",
    "Estimated Value", "Probability %", "Expected Close Date",
    "Assigned To", "Date Created", "Last Contact", "Next Action",
    "Next Action Date", "Notes"
]

# Style for headers
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True, size=11)
header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Add headers
for col_num, header in enumerate(leads_headers, 1):
    cell = ws_leads.cell(row=1, column=col_num)
    cell.value = header
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = header_alignment

# Sample lead data
leads_data = [
    [1, "Acme Corporation", "John Smith", "VP of Sales", "john.smith@acme.com", "+1-555-0101",
     "Technology", "USA", "Website", "Qualified", "High", 75000, 60,
     datetime(2025, 11, 15), "Alice Johnson", datetime(2025, 10, 1), datetime(2025, 10, 22),
     "Send proposal", datetime(2025, 10, 25), "Interested in enterprise plan with 50 licenses"],

    [2, "TechStart Inc", "Sarah Johnson", "CTO", "sarah@techstart.com", "+1-555-0102",
     "SaaS", "USA", "Referral", "Proposal Sent", "High", 125000, 75,
     datetime(2025, 11, 30), "Bob Wilson", datetime(2025, 9, 15), datetime(2025, 10, 23),
     "Follow-up call", datetime(2025, 10, 26), "Requested custom integration features"],

    [3, "Global Solutions Ltd", "Mike Chen", "Director of Operations", "mchen@globalsol.com", "+44-20-5555-0103",
     "Manufacturing", "UK", "LinkedIn", "Negotiation", "High", 200000, 80,
     datetime(2025, 12, 10), "Alice Johnson", datetime(2025, 9, 20), datetime(2025, 10, 23),
     "Contract review", datetime(2025, 10, 27), "Legal team reviewing contract terms"],

    [4, "SmartBiz LLC", "Emily Davis", "Marketing Manager", "emily@smartbiz.com", "+1-555-0104",
     "Retail", "USA", "Trade Show", "Nurturing", "Medium", 35000, 30,
     datetime(2026, 1, 20), "Carol Martinez", datetime(2025, 10, 10), datetime(2025, 10, 20),
     "Send case study", datetime(2025, 10, 28), "Budget available Q1 2026"],

    [5, "Innovation Labs", "David Park", "CEO", "dpark@innovlabs.com", "+1-555-0105",
     "Technology", "USA", "Cold Call", "Demo Scheduled", "High", 95000, 50,
     datetime(2025, 11, 20), "Bob Wilson", datetime(2025, 10, 18), datetime(2025, 10, 23),
     "Product demo", datetime(2025, 10, 28), "Requested technical demo for engineering team"],

    [6, "Enterprise Systems Co", "Lisa Anderson", "Procurement Officer", "lisa@enterprisesys.com", "+1-555-0106",
     "Finance", "USA", "Partner", "Qualified", "Medium", 85000, 55,
     datetime(2025, 12, 5), "Alice Johnson", datetime(2025, 10, 5), datetime(2025, 10, 22),
     "Send ROI analysis", datetime(2025, 10, 29), "Comparing with competitors"],

    [7, "Digital Ventures", "Tom Brown", "Product Manager", "tom@digitalventures.com", "+1-555-0107",
     "Media", "Canada", "Website", "New", "Low", 20000, 20,
     datetime(2026, 2, 1), "Carol Martinez", datetime(2025, 10, 22), datetime(2025, 10, 22),
     "Initial call", datetime(2025, 10, 30), "Submitted contact form on website"],

    [8, "HealthTech Solutions", "Maria Garcia", "COO", "maria@healthtech.com", "+1-555-0108",
     "Healthcare", "USA", "Referral", "Proposal Sent", "High", 150000, 70,
     datetime(2025, 11, 25), "Bob Wilson", datetime(2025, 9, 25), datetime(2025, 10, 21),
     "Schedule stakeholder meeting", datetime(2025, 10, 31), "Requires HIPAA compliance features"],

    [9, "Green Energy Corp", "James Wilson", "VP of Technology", "james@greenenergy.com", "+1-555-0109",
     "Energy", "USA", "Conference", "Qualified", "Medium", 60000, 45,
     datetime(2025, 12, 15), "Alice Johnson", datetime(2025, 10, 12), datetime(2025, 10, 23),
     "Send pricing options", datetime(2025, 11, 1), "Met at industry conference"],

    [10, "Retail Plus Group", "Jennifer Lee", "IT Director", "jennifer@retailplus.com", "+1-555-0110",
     "Retail", "USA", "Cold Call", "Contacted", "Low", 40000, 25,
     datetime(2026, 1, 10), "Carol Martinez", datetime(2025, 10, 19), datetime(2025, 10, 23),
     "Follow-up email", datetime(2025, 11, 5), "Initial interest but limited budget"],
]

# Add data to Leads sheet
for row_num, lead in enumerate(leads_data, 2):
    for col_num, value in enumerate(lead, 1):
        cell = ws_leads.cell(row=row_num, column=col_num)
        cell.value = value

        # Format currency columns
        if col_num == 12:  # Estimated Value
            cell.number_format = '$#,##0'

        # Format percentage columns
        if col_num == 13:  # Probability
            cell.number_format = '0"%"'

        # Format date columns
        if col_num in [14, 16, 17, 19]:  # Date columns
            cell.number_format = 'yyyy-mm-dd'

        # Add borders
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        cell.border = thin_border

# Auto-adjust column widths
for col in range(1, len(leads_headers) + 1):
    ws_leads.column_dimensions[get_column_letter(col)].width = 18

# Freeze top row
ws_leads.freeze_panes = "A2"

# ===== ACTIVITIES SHEET =====
ws_activities = wb.create_sheet("Activities", 1)

activities_headers = [
    "Activity ID", "Lead ID", "Company Name", "Activity Type",
    "Activity Date", "Performed By", "Duration (min)", "Outcome", "Notes"
]

for col_num, header in enumerate(activities_headers, 1):
    cell = ws_activities.cell(row=1, column=col_num)
    cell.value = header
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = header_alignment

activities_data = [
    [1, 1, "Acme Corporation", "Phone Call", datetime(2025, 10, 22), "Alice Johnson", 30, "Positive", "Discussed requirements and timeline"],
    [2, 2, "TechStart Inc", "Email", datetime(2025, 10, 23), "Bob Wilson", 15, "Proposal Sent", "Sent detailed proposal with pricing"],
    [3, 3, "Global Solutions Ltd", "Meeting", datetime(2025, 10, 23), "Alice Johnson", 60, "Positive", "Contract negotiation meeting with legal team"],
    [4, 4, "SmartBiz LLC", "Email", datetime(2025, 10, 20), "Carol Martinez", 10, "Pending", "Sent product brochure"],
    [5, 5, "Innovation Labs", "Phone Call", datetime(2025, 10, 23), "Bob Wilson", 25, "Demo Scheduled", "Scheduled demo for October 28"],
    [6, 6, "Enterprise Systems Co", "Meeting", datetime(2025, 10, 22), "Alice Johnson", 45, "Positive", "ROI discussion and case studies"],
    [7, 8, "HealthTech Solutions", "Phone Call", datetime(2025, 10, 21), "Bob Wilson", 40, "Positive", "Discussed HIPAA compliance requirements"],
    [8, 9, "Green Energy Corp", "Email", datetime(2025, 10, 23), "Alice Johnson", 15, "Pending", "Sent pricing for different tiers"],
    [9, 10, "Retail Plus Group", "Phone Call", datetime(2025, 10, 23), "Carol Martinez", 20, "Follow-up Needed", "Budget constraints discussed"],
    [10, 1, "Acme Corporation", "Email", datetime(2025, 10, 15), "Alice Johnson", 10, "Positive", "Initial contact and qualification"],
]

for row_num, activity in enumerate(activities_data, 2):
    for col_num, value in enumerate(activity, 1):
        cell = ws_activities.cell(row=row_num, column=col_num)
        cell.value = value

        if col_num == 5:  # Date column
            cell.number_format = 'yyyy-mm-dd hh:mm'

        cell.border = thin_border

for col in range(1, len(activities_headers) + 1):
    ws_activities.column_dimensions[get_column_letter(col)].width = 20

ws_activities.freeze_panes = "A2"

# ===== PIPELINE SHEET =====
ws_pipeline = wb.create_sheet("Pipeline Summary", 2)

pipeline_headers = ["Status", "Count", "Total Value", "Avg Deal Size", "Weighted Value"]

for col_num, header in enumerate(pipeline_headers, 1):
    cell = ws_pipeline.cell(row=1, column=col_num)
    cell.value = header
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = header_alignment

# Pipeline summary data (calculated from leads)
status_summary = {}
for lead in leads_data:
    status = lead[9]  # Status column
    value = lead[11]  # Estimated Value
    probability = lead[12]  # Probability

    if status not in status_summary:
        status_summary[status] = {'count': 0, 'total_value': 0, 'weighted_value': 0}

    status_summary[status]['count'] += 1
    status_summary[status]['total_value'] += value
    status_summary[status]['weighted_value'] += (value * probability / 100)

row_num = 2
for status, data in status_summary.items():
    ws_pipeline.cell(row=row_num, column=1).value = status
    ws_pipeline.cell(row=row_num, column=2).value = data['count']
    ws_pipeline.cell(row=row_num, column=3).value = data['total_value']
    ws_pipeline.cell(row=row_num, column=3).number_format = '$#,##0'
    ws_pipeline.cell(row=row_num, column=4).value = data['total_value'] / data['count']
    ws_pipeline.cell(row=row_num, column=4).number_format = '$#,##0'
    ws_pipeline.cell(row=row_num, column=5).value = data['weighted_value']
    ws_pipeline.cell(row=row_num, column=5).number_format = '$#,##0'

    for col in range(1, 6):
        ws_pipeline.cell(row=row_num, column=col).border = thin_border

    row_num += 1

for col in range(1, len(pipeline_headers) + 1):
    ws_pipeline.column_dimensions[get_column_letter(col)].width = 20

# ===== CONTACTS SHEET =====
ws_contacts = wb.create_sheet("Contacts", 3)

contacts_headers = [
    "Contact ID", "Full Name", "Title", "Company", "Email",
    "Phone", "Mobile", "LinkedIn", "Last Contact", "Relationship"
]

for col_num, header in enumerate(contacts_headers, 1):
    cell = ws_contacts.cell(row=1, column=col_num)
    cell.value = header
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = header_alignment

contacts_data = [
    [1, "John Smith", "VP of Sales", "Acme Corporation", "john.smith@acme.com",
     "+1-555-0101", "+1-555-0201", "linkedin.com/in/johnsmith", datetime(2025, 10, 22), "Warm"],
    [2, "Sarah Johnson", "CTO", "TechStart Inc", "sarah@techstart.com",
     "+1-555-0102", "+1-555-0202", "linkedin.com/in/sarahjohnson", datetime(2025, 10, 23), "Hot"],
    [3, "Mike Chen", "Director of Operations", "Global Solutions Ltd", "mchen@globalsol.com",
     "+44-20-5555-0103", "+44-77-5555-0203", "linkedin.com/in/mikechen", datetime(2025, 10, 23), "Hot"],
    [4, "Emily Davis", "Marketing Manager", "SmartBiz LLC", "emily@smartbiz.com",
     "+1-555-0104", "+1-555-0204", "linkedin.com/in/emilydavis", datetime(2025, 10, 20), "Cold"],
    [5, "David Park", "CEO", "Innovation Labs", "dpark@innovlabs.com",
     "+1-555-0105", "+1-555-0205", "linkedin.com/in/davidpark", datetime(2025, 10, 23), "Warm"],
]

for row_num, contact in enumerate(contacts_data, 2):
    for col_num, value in enumerate(contact, 1):
        cell = ws_contacts.cell(row=row_num, column=col_num)
        cell.value = value

        if col_num == 9:  # Last Contact date
            cell.number_format = 'yyyy-mm-dd'

        cell.border = thin_border

for col in range(1, len(contacts_headers) + 1):
    ws_contacts.column_dimensions[get_column_letter(col)].width = 22

ws_contacts.freeze_panes = "A2"

# ===== DASHBOARD SHEET =====
ws_dashboard = wb.create_sheet("Dashboard", 4)

# Title
ws_dashboard.merge_cells('A1:F1')
title_cell = ws_dashboard['A1']
title_cell.value = "LEAD MANAGEMENT DASHBOARD"
title_cell.font = Font(size=20, bold=True, color="4472C4")
title_cell.alignment = Alignment(horizontal="center", vertical="center")

# Key Metrics
ws_dashboard['A3'] = "KEY METRICS"
ws_dashboard['A3'].font = Font(size=14, bold=True)

metrics_headers = ["Metric", "Value"]
for col_num, header in enumerate(metrics_headers, 1):
    cell = ws_dashboard.cell(row=4, column=col_num)
    cell.value = header
    cell.fill = header_fill
    cell.font = header_font

# Calculate metrics
total_leads = len(leads_data)
total_value = sum(lead[11] for lead in leads_data)
avg_deal_size = total_value / total_leads
weighted_pipeline = sum(lead[11] * lead[12] / 100 for lead in leads_data)
high_priority_count = sum(1 for lead in leads_data if lead[10] == "High")

metrics = [
    ["Total Leads", total_leads],
    ["Total Pipeline Value", f"${total_value:,}"],
    ["Average Deal Size", f"${avg_deal_size:,.0f}"],
    ["Weighted Pipeline", f"${weighted_pipeline:,.0f}"],
    ["High Priority Leads", high_priority_count],
]

for row_num, metric in enumerate(metrics, 5):
    ws_dashboard.cell(row=row_num, column=1).value = metric[0]
    ws_dashboard.cell(row=row_num, column=2).value = metric[1]
    ws_dashboard.cell(row=row_num, column=1).font = Font(bold=True)

# Sales Rep Performance
ws_dashboard['D3'] = "SALES REP PERFORMANCE"
ws_dashboard['D3'].font = Font(size=14, bold=True)

rep_headers = ["Sales Rep", "Leads", "Total Value"]
for col_num, header in enumerate(rep_headers, 4):
    cell = ws_dashboard.cell(row=4, column=col_num)
    cell.value = header
    cell.fill = header_fill
    cell.font = header_font

# Calculate rep performance
rep_performance = {}
for lead in leads_data:
    rep = lead[14]  # Assigned To
    value = lead[11]  # Estimated Value

    if rep not in rep_performance:
        rep_performance[rep] = {'count': 0, 'value': 0}

    rep_performance[rep]['count'] += 1
    rep_performance[rep]['value'] += value

row_num = 5
for rep, data in rep_performance.items():
    ws_dashboard.cell(row=row_num, column=4).value = rep
    ws_dashboard.cell(row=row_num, column=5).value = data['count']
    ws_dashboard.cell(row=row_num, column=6).value = f"${data['value']:,}"
    row_num += 1

# Adjust column widths
for col in range(1, 7):
    ws_dashboard.column_dimensions[get_column_letter(col)].width = 22

# Save the workbook
wb.save("C:\\Users\\nitza\\devprojects\\personal-assistant\\Lead_Management_System.xlsx")
print("Advanced Lead Management System created successfully!")
print("File: Lead_Management_System.xlsx")
print("\nSheets included:")
print("1. Leads - Main lead tracking with 10 sample leads")
print("2. Activities - Track all interactions and follow-ups")
print("3. Pipeline Summary - Status breakdown with metrics")
print("4. Contacts - Contact information database")
print("5. Dashboard - Key metrics and sales rep performance")
