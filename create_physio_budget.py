# ABOUTME: Creates comprehensive budget spreadsheet for physiotherapy practice
# ABOUTME: Includes budget overview, income tracking, expenses, and client management

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime
import os

def create_physio_budget():
    """Create comprehensive physiotherapy practice budget spreadsheet"""

    # Create workbook
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Color scheme
    HEADER_FILL = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    INPUT_FILL = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
    TOTAL_FILL = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
    INCOME_FILL = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
    EXPENSE_FILL = PatternFill(start_color='E74C3C', end_color='E74C3C', fill_type='solid')

    HEADER_FONT = Font(bold=True, color='FFFFFF', size=11)
    BOLD_FONT = Font(bold=True)
    THIN_BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # === BUDGET OVERVIEW SHEET ===
    ws_budget = wb.create_sheet('Budget Overview', 0)
    ws_budget.column_dimensions['A'].width = 20
    for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M']:
        ws_budget.column_dimensions[col].width = 12

    # Title
    ws_budget['A1'] = 'Physiotherapy Practice - Budget Overview'
    ws_budget['A1'].font = Font(bold=True, size=14)
    ws_budget.merge_cells('A1:M1')

    ws_budget['A2'] = 'Year: 2025'
    ws_budget['A2'].font = BOLD_FONT

    # Monthly headers
    months = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
    ws_budget['A4'] = 'Category'
    for i, month in enumerate(months):
        cell = ws_budget.cell(4, i+2)
        cell.value = month
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center')

    # Income section
    ws_budget['A5'] = 'INCOME'
    ws_budget['A5'].font = BOLD_FONT
    ws_budget['A5'].fill = INCOME_FILL

    income_categories = ['Session Fees', 'Home Visits', 'Group Classes', 'Other Income']
    row = 6
    for cat in income_categories:
        ws_budget.cell(row, 1).value = cat
        for col in range(2, 14):
            cell = ws_budget.cell(row, col)
            cell.value = 0
            cell.number_format = '#,##0'
            cell.border = THIN_BORDER
        row += 1

    ws_budget.cell(row, 1).value = 'Total Income'
    ws_budget.cell(row, 1).font = BOLD_FONT
    for col in range(2, 14):
        cell = ws_budget.cell(row, col)
        cell.value = f'=SUM({get_column_letter(col)}6:{get_column_letter(col)}{row-1})'
        cell.fill = TOTAL_FILL
        cell.font = BOLD_FONT
        cell.number_format = '#,##0'
        cell.border = THIN_BORDER

    total_income_row = row

    # Expense section
    expense_start_row = row + 2
    ws_budget.cell(expense_start_row, 1).value = 'EXPENSES'
    ws_budget.cell(expense_start_row, 1).font = BOLD_FONT
    ws_budget.cell(expense_start_row, 1).fill = EXPENSE_FILL

    expense_categories = ['Rent/Utilities', 'Equipment & Supplies', 'Professional Development',
                          'Insurance', 'Marketing', 'Transportation', 'Administrative', 'Other Expenses']
    row = expense_start_row + 1
    for cat in expense_categories:
        ws_budget.cell(row, 1).value = cat
        for col in range(2, 14):
            cell = ws_budget.cell(row, col)
            cell.value = 0
            cell.number_format = '#,##0'
            cell.border = THIN_BORDER
        row += 1

    ws_budget.cell(row, 1).value = 'Total Expenses'
    ws_budget.cell(row, 1).font = BOLD_FONT
    for col in range(2, 14):
        cell = ws_budget.cell(row, col)
        cell.value = f'=SUM({get_column_letter(col)}{expense_start_row+1}:{get_column_letter(col)}{row-1})'
        cell.fill = TOTAL_FILL
        cell.font = BOLD_FONT
        cell.number_format = '#,##0'
        cell.border = THIN_BORDER

    total_expense_row = row

    # Net Income
    row += 2
    ws_budget.cell(row, 1).value = 'NET INCOME'
    ws_budget.cell(row, 1).font = Font(bold=True, size=12)
    for col in range(2, 14):
        cell = ws_budget.cell(row, col)
        cell.value = f'={get_column_letter(col)}{total_income_row}-{get_column_letter(col)}{total_expense_row}'
        cell.fill = INCOME_FILL
        cell.font = BOLD_FONT
        cell.number_format = '#,##0'
        cell.border = THIN_BORDER

    # === INCOME TRACKER SHEET ===
    ws_income = wb.create_sheet('Income Tracker', 1)

    widths = {'A': 12, 'B': 20, 'C': 25, 'D': 20, 'E': 12, 'F': 12, 'G': 25}
    for col, width in widths.items():
        ws_income.column_dimensions[col].width = width

    ws_income['A1'] = 'Income Tracker - 2025'
    ws_income['A1'].font = Font(bold=True, size=14)
    ws_income.merge_cells('A1:G1')

    headers = ['Date', 'Client Name', 'Service Type', 'Payment Method', 'Amount', 'Status', 'Notes']
    for col, header in enumerate(headers, 1):
        cell = ws_income.cell(3, col)
        cell.value = header
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center')
        cell.border = THIN_BORDER

    # Add data validation dropdowns
    service_dv = DataValidation(type="list", formula1='"Regular Session,Home Visit,Initial Consultation,Group Class,Follow-up,Other"')
    ws_income.add_data_validation(service_dv)
    service_dv.add(f'C4:C203')

    payment_dv = DataValidation(type="list", formula1='"Cash,Credit Card,Bank Transfer,Insurance,Check"')
    ws_income.add_data_validation(payment_dv)
    payment_dv.add(f'D4:D203')

    status_dv = DataValidation(type="list", formula1='"Paid,Pending,Partial"')
    ws_income.add_data_validation(status_dv)
    status_dv.add(f'F4:F203')

    # Format data rows
    for row in range(4, 204):
        ws_income.cell(row, 1).number_format = 'DD/MM/YYYY'
        ws_income.cell(row, 5).number_format = '#,##0'
        for col in range(1, 8):
            ws_income.cell(row, col).border = THIN_BORDER

    # Summary
    summary_row = 205
    ws_income.cell(summary_row, 4).value = 'TOTAL INCOME:'
    ws_income.cell(summary_row, 4).font = BOLD_FONT
    ws_income.cell(summary_row, 5).value = f'=SUM(E4:E203)'
    ws_income.cell(summary_row, 5).font = BOLD_FONT
    ws_income.cell(summary_row, 5).number_format = '#,##0'
    ws_income.cell(summary_row, 5).fill = TOTAL_FILL

    ws_income.cell(summary_row+1, 4).value = 'Paid:'
    ws_income.cell(summary_row+1, 5).value = '=SUMIF(F4:F203,"Paid",E4:E203)'
    ws_income.cell(summary_row+1, 5).number_format = '#,##0'

    ws_income.cell(summary_row+2, 4).value = 'Pending:'
    ws_income.cell(summary_row+2, 5).value = '=SUMIF(F4:F203,"Pending",E4:E203)'
    ws_income.cell(summary_row+2, 5).number_format = '#,##0'

    # === EXPENSE TRACKER SHEET ===
    ws_expense = wb.create_sheet('Expense Tracker', 2)

    widths = {'A': 12, 'B': 25, 'C': 20, 'D': 12, 'E': 25}
    for col, width in widths.items():
        ws_expense.column_dimensions[col].width = width

    ws_expense['A1'] = 'Expense Tracker - 2025'
    ws_expense['A1'].font = Font(bold=True, size=14)
    ws_expense.merge_cells('A1:E1')

    headers = ['Date', 'Category', 'Description', 'Amount', 'Notes']
    for col, header in enumerate(headers, 1):
        cell = ws_expense.cell(3, col)
        cell.value = header
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center')
        cell.border = THIN_BORDER

    # Category dropdown
    category_dv = DataValidation(type="list", formula1='"Rent/Utilities,Equipment & Supplies,Professional Development,Insurance,Marketing,Transportation,Administrative,Other Expenses"')
    ws_expense.add_data_validation(category_dv)
    category_dv.add(f'B4:B203')

    # Format data rows
    for row in range(4, 204):
        ws_expense.cell(row, 1).number_format = 'DD/MM/YYYY'
        ws_expense.cell(row, 4).number_format = '#,##0'
        for col in range(1, 6):
            ws_expense.cell(row, col).border = THIN_BORDER

    # Summary
    summary_row = 205
    ws_expense.cell(summary_row, 3).value = 'TOTAL EXPENSES:'
    ws_expense.cell(summary_row, 3).font = BOLD_FONT
    ws_expense.cell(summary_row, 4).value = f'=SUM(D4:D203)'
    ws_expense.cell(summary_row, 4).font = BOLD_FONT
    ws_expense.cell(summary_row, 4).number_format = '#,##0'
    ws_expense.cell(summary_row, 4).fill = TOTAL_FILL

    # Category breakdown
    ws_expense.cell(summary_row+2, 2).value = 'Expense Breakdown:'
    ws_expense.cell(summary_row+2, 2).font = BOLD_FONT

    expense_cats = ['Rent/Utilities', 'Equipment & Supplies', 'Professional Development',
                    'Insurance', 'Marketing', 'Transportation', 'Administrative', 'Other Expenses']
    for i, cat in enumerate(expense_cats):
        row_num = summary_row + 3 + i
        ws_expense.cell(row_num, 2).value = cat
        ws_expense.cell(row_num, 4).value = f'=SUMIF(B4:B203,"{cat}",D4:D203)'
        ws_expense.cell(row_num, 4).number_format = '#,##0'

    # === CLIENT MANAGEMENT SHEET ===
    ws_clients = wb.create_sheet('Client Management', 3)

    widths = {'A': 8, 'B': 20, 'C': 15, 'D': 25, 'E': 12, 'F': 15, 'G': 20, 'H': 25}
    for col, width in widths.items():
        ws_clients.column_dimensions[col].width = width

    ws_clients['A1'] = 'Client Management Database'
    ws_clients['A1'].font = Font(bold=True, size=14)
    ws_clients.merge_cells('A1:H1')

    headers = ['ID', 'Client Name', 'Phone', 'Email', 'First Visit', 'Last Visit', 'Total Sessions', 'Notes']
    for col, header in enumerate(headers, 1):
        cell = ws_clients.cell(3, col)
        cell.value = header
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center')
        cell.border = THIN_BORDER

    # Format data rows
    for row in range(4, 104):
        ws_clients.cell(row, 1).value = row - 3  # Auto ID
        ws_clients.cell(row, 5).number_format = 'DD/MM/YYYY'
        ws_clients.cell(row, 6).number_format = 'DD/MM/YYYY'
        for col in range(1, 9):
            ws_clients.cell(row, col).border = THIN_BORDER

    # Instructions sheet
    ws_instructions = wb.create_sheet('Instructions', 4)
    ws_instructions.column_dimensions['A'].width = 80

    instructions = [
        "Physiotherapy Practice Budget Manager - Instructions",
        "",
        "BUDGET OVERVIEW:",
        "- Monthly summary of income and expenses",
        "- Update monthly totals based on Income/Expense Tracker sheets",
        "- Green shows net income (profit/loss) per month",
        "",
        "INCOME TRACKER:",
        "- Record each client session/payment",
        "- Use dropdowns for Service Type, Payment Method, and Status",
        "- Amount column calculates totals automatically",
        "- Track pending payments vs received",
        "",
        "EXPENSE TRACKER:",
        "- Record all business expenses",
        "- Use Category dropdown to match budget categories",
        "- Summary shows total expenses and breakdown by category",
        "",
        "CLIENT MANAGEMENT:",
        "- Maintain client contact database",
        "- Track first visit, last visit, and total sessions",
        "- Use for client history and follow-ups",
        "",
        "TIPS:",
        "- Update Income/Expense trackers daily or weekly",
        "- Review Budget Overview monthly",
        "- Use Status field in Income Tracker to track unpaid sessions",
        "- Keep client notes for medical history or preferences"
    ]

    for i, line in enumerate(instructions, 1):
        cell = ws_instructions.cell(i, 1)
        cell.value = line
        if i == 1:
            cell.font = Font(bold=True, size=14)
        elif line.endswith(':'):
            cell.font = BOLD_FONT

    # Save workbook
    output_path = os.path.join(os.getcwd(), 'Physiotherapy_Budget_2025.xlsx')
    wb.save(output_path)
    print(f'[OK] Created: {output_path}')
    return output_path

if __name__ == '__main__':
    create_physio_budget()
