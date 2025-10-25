from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Create a new workbook
wb = Workbook()
ws = wb.active
ws.title = "Weekly Tasks"

# Define styles
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True, size=12)
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Set column widths
ws.column_dimensions['A'].width = 40
ws.column_dimensions['B'].width = 15
ws.column_dimensions['C'].width = 15
ws.column_dimensions['D'].width = 15
ws.column_dimensions['E'].width = 30

# Create headers
headers = ['Task', 'Priority', 'Status', 'Due Date', 'Notes']
for col_num, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_num)
    cell.value = header
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = border

# Add sample data structure with days of the week
days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
row = 2

for day in days:
    # Day header
    cell = ws.cell(row=row, column=1)
    cell.value = day
    cell.font = Font(bold=True, size=11)
    cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    ws.merge_cells(f'A{row}:E{row}')
    cell.alignment = Alignment(horizontal='left', vertical='center')
    cell.border = border
    row += 1

    # Add 3 task rows per day
    for i in range(3):
        for col in range(1, 6):
            cell = ws.cell(row=row, column=col)
            cell.border = border
            cell.alignment = Alignment(vertical='center')

            # Add dropdown values hint in first row
            if row == 3 and col == 2:
                cell.value = "High/Medium/Low"
            elif row == 3 and col == 3:
                cell.value = "Not Started/In Progress/Done"
        row += 1

# Freeze the header row
ws.freeze_panes = 'A2'

# Save the workbook
wb.save('Weekly_Task_Tracker.xlsx')
print("Weekly task tracker created successfully!")
